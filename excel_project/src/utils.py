"""
File: utils.py
Author: Kyle Fullerton
Purpose: File to put any miscellaneous functions that can be helpful
in other places in the program.
"""
from openpyxl.styles import (Color,
                             Border,
                             Side,
                             Alignment,
                             PatternFill
                             )
import sys

"""
Method: add_header
Purpose: Handles all of the stlying for the header cells.
Also, inserts the header title as well.

Parameters: 
cell- cell object
value - header string to be inserted
"""


def add_header(cell, value):
    gray = Color(rgb='00C2C2C2')
    black = Color(rgb='000000')

    border = Border(left=Side(border_style='thin', color=black),
                    right=Side(border_style='thin', color=black),
                    top=Side(border_style='thin', color=black),
                    bottom=Side(border_style='thin', color=black))

    cell.fill = PatternFill(patternType='solid', fgColor=gray)
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
    cell.value = value


"""
Method: add_assembly_num
Purpose: Adds a new header to the spreadsheet 
and to the header list.

Parameters: 
header_list- list of headers for the spreadsheet
write_sheet- spreadsheet that is being written to
read_sheet- spreadsheet to read the assembly number
configs- dictionary of configuration parameters
"""


def add_assembly_num(header_list, write_sheet, read_sheet, configs):
    try:
        assembly_num = read_sheet.cell_value(configs["part_num_row"] - 1,
                                             configs["part_num_column"] - 1)

    except IndexError:
        print("Error: config parameter {0} or {1} defines an out of range row/column"
              .format("'part_num_row'", "'part_num_column'"))
        sys.exit(1)

    if assembly_num not in header_list:
        add_header(write_sheet.cell(configs["header_row"], len(header_list) + 1), assembly_num)
        header_list.append(assembly_num)


"""
Method: edit_column_width
Purpose: Changes the dimensions of all the columns
used in the spreadsheet to the specified width.

Parameters: 
out_write_sheet- spreadsheet that is written to
headers- list of headers
configs- dictionary of configuration parameters
"""


def edit_column_width(out_write_sheet, headers, configs):
    for i in range(0, len(headers)):
        column = get_column_letter(i)
        out_write_sheet.column_dimensions[column].width = configs["column_width"]


"""
Method: get_column_num
Purpose: Utility function to get a column number if a
letter may be specified instead.

Parameter: 
number- string or int that will be converted to a string

Return: 
col_num- Column number corresponding to its column letter.
Ex: 2 = 'C'.
"""


def get_column_num(letters):
    try:
        column_num = int(letters)

    except ValueError:
        letters_str = letters
        column_num = 0

        while len(letters_str) > 1:
            column_num += 26
            letters_str = letters_str[1:]
        column_num += ord(letters[0].upper()) - ord("A")

    return column_num


"""
Method: get_column_letter
Purpose: Utility function to get a column letter if a
number may be specified instead.

Parameter: 
number- string or int that will be converted to a string
Return: 
col_ltr- Column letter corresponding to its column number.
Ex: 'C' = 2.
"""


def get_column_letter(number):
    try:
        column_num = int(number)

        column_letter = ""
        while column_num > 25:
            column_num -= 26
            column_letter += "A"
        column_letter += chr(column_num + ord("A"))

    except ValueError:
        column_letter = number

    return column_letter


"""
Method: get_range
Purpose: To construct a range of cells as a string.
Ex: "A1:D1"

Parameters:
column1- column for the first cell in the range
column2- column for the second cell in the range
row1- row for the first cell in the range
row2- row for the second cell in the range 

Return:
A string that represents a range of cells.
"""


def get_range(column1, column2, row1, row2):
    start = str(get_column_letter(column1)) + str(row1)
    end = str(get_column_letter(column2)) + str(row2)

    return start + ":" + end
