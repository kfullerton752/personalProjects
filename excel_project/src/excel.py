"""
File: excel.py
Author: Kyle Fullerton
Purpose: File that includes functions pertaining to updating
the output excel spreadsheet.
"""

import sys
from excelScript import utils
from openpyxl.styles import Alignment


"""
Method: update_master
Purpose: Grabs the column of serial numbers and loops through every row in
the spreadsheet starting with the first place that wanted data appears in 
the spreadsheet (Like you only care about rows 9 to n rows for example). 
If the serial number is in the dictionary that represents the current output
file then that entry is checked if additional information needs to be added
to remarks. Tne entry is also checked to see if a quantity needs to be added,
update, or replaced. If the serial number isn't in the dictionary, then 
a new line is appended to the spreadsheet and a new entry is added into
the dictionary. 

Parameters: read_sheet- spreadsheet used to read values from
write_sheet- output spreadsheet we are writing to
header_list- list of headers for the spreadsheet
part_dict- dictionary of values from the output excel file
config_dict- dictionary of configuration parameters
file_name- name of the current file being read in

Variables:
lines_skipped- number of lines skipped when reading the input excel file
assembly_num- assembly number for the read in excel sheet
column_num- column number corresponding to the serial number in the input 
excel file
serial_nums- serial numbers found in the serial number column
row_data- all of the data in the current row of the input spreadsheet
part_num- current part number in the row
column- column where the assembly number is found
part_num_row- row number for the current row

Return- part_dict- updated dictionary of values from the xlw sheet
"""


def update_master(read_sheet, write_sheet, header_list, part_dict,
                  config_dict, file_name, workbook):

    lines_skipped = 0
    assembly_num = read_sheet.cell_value(config_dict["part_num_row"] - 1,
                                         config_dict["part_num_column"] - 1)
    column_num = utils.get_column_num(config_dict["serial_num_column"])

    try:
        serial_nums = read_sheet.col_values(column_num)

    except IndexError:
        print("Error: config parameter {0} defines an out of range column"
              "for the sheet in file {1}".format("'serial_num_column'", file_name))
        return part_dict

    # loops through the read_sheet from where we care about the data
    for row in range(config_dict["data_start"] - 1, len(serial_nums)):
        row_data = pull_data(row, read_sheet, config_dict)

        # got a row_data with not all of the input needed
        if len(row_data) == 0:
            lines_skipped += 1
            continue

        # if part_num already in dictionary then just add a qty to the respective
        # assembly number
        part_num = str(row_data[0]).strip()
        row_data[1] = row_data[1].upper()

        column = header_list.index(assembly_num)

        if part_num in part_dict:

            part_num_row = part_dict[part_num]
            update_remarks(part_num_row, config_dict, row_data, write_sheet)

            # update the qty for the card appropriately
            if len(part_num_row) == len(header_list):

                if part_num_row[column][0] == "":
                    part_num_row[column][0] = row_data[-1]

                else:
                    # either adds or replaces qty depending on specified mode
                    if config_dict["add_mode"]:
                        part_num_row[column][0] += row_data[-1]
                    else:
                        part_num_row[column][0] = row_data[-1]

                write_sheet.cell(part_num_row[0], column + 1).value = part_num_row[column][0]

            # add new column to the id's row_data
            else:
                part_num_row.append([row_data[-1], column + 1])
                write_sheet.cell(part_num_row[0], column + 1).value = row_data[-1]

            part_dict[part_num] = part_num_row
            update_totals(part_num_row[0], part_num, config_dict, workbook, header_list)

        else:
            # add row_data to spreadsheet
            write_sheet.append(row_data[:-1])

            # add qty to spreadsheet
            row_data[0] = len(part_dict) + config_dict["header_row"] + 1
            write_sheet.cell(row_data[0], column + 1).value = row_data[-1]

            # add new entry to dictionary
            row_data[-1] = [row_data[-1], column + 1]
            part_dict[part_num] = row_data

            update_totals(row_data[0], part_num, config_dict, workbook, header_list)
    if config_dict["lines_skipped"]:
        print("Number of lines skipped in file {0}: {1}".format(file_name, lines_skipped))

    return part_dict


"""
Method: update_remarks
Purpose: Updates the remarks column if a different company
is found for the same part number. 

Parameters:
part_num_row- row where the part number is located
config_dict- dictionary of configuration parameters
row_data- all of the data in the current row
write_sheet- worksheet for the output excel file

Variable- remarks- column where the remarks column is located
"""


def update_remarks(part_num_row, config_dict, row_data, write_sheet):
    remarks = config_dict["out_remarks_index"] - 1
    if remarks >= len(part_num_row) or remarks >= len(row_data):
        print("Error: config parameter {0} defines out of range column for sheet"
              .format("'out_remarks_index'"))
        sys.exit(1)

    # adds additional company for remarks if needed
    if row_data[remarks] not in part_num_row[remarks]:
        part_num_row[remarks] = part_num_row[remarks] + "/" + row_data[remarks]

        write_sheet.cell(part_num_row[0], remarks + 1).value = part_num_row[remarks]
        write_sheet.cell(part_num_row[0], remarks + 1).alignment = Alignment(wrap_text=True)


"""
Method: pull_data
Purpose: Checks that all of the specified columns have
data. Then creates and returns a list of all data
from the specified columns from the read in spreadsheet.

Parameters:
row- current row in the spreadsheet
sheet- spreadsheet that the function is reading from
config_dict- dictionary of configuration options

Return: row_data- list of all of the specified data
in the current row
"""


def pull_data(row, sheet, config_dict):
    row_data = []

    for column in config_dict["check_columns"]:
        try:
            if sheet.cell_value(row, column - 1) == "":
                return row_data
        except IndexError:
            return row_data

    for column in config_dict["wanted_columns"]:
        try:
            row_data.append(sheet.cell_value(row, int(column) - 1))

        except IndexError:
            continue

    return row_data


"""
Method: add_headers
Purpose: Gets the initial headers from the sheet.
If no headers exist then the specified 
headers are added to the sheet.

Parameters: out_read_sheet- read excel sheet
out_write_sheet- write excel sheet
configs- dictionary of configuration parameters
"""


def add_headers(out_read_sheet, out_write_sheet, configs):
    try:
        header_list = out_read_sheet.row_values(configs["header_row"] - 1)

    except IndexError:
        header_list = configs["out_default_headers"]
        for i in range(0, len(header_list)):
            utils.add_header(out_write_sheet.cell(configs["header_row"], i + 1), header_list[i])

    return header_list


"""
Method: update_totals
Purpose: Used to create a totals sheet for the total
number of parts needed to make a specified number 
of cards. First, checks if a new sheet needs to be
added. Then adds the headers to the total sheet. 
Next, constructs a SUMPRODUCT formula for the passed
in row. This formula and the part number are added 
to the total sheet.
Note: SUMPRODUCT takes two plus equal length arrays
from ranges of cells, multiplies the same index
in each array by each other, and then sums the result.
Ex: SUMPRODUCT([1,2,3], [4, 5, 6]) = 4 + 10 + 18 = 32

Parameters:
row- row number from the first sheet
part_num- part number 
config_dict- dictionary of configuration parameters 
workbook- workbook object from the output excel file
header_list- list of headers on the output excel spreadsheet


Variables:
title- title of the specified spreadsheet
sheet_headers- headers wanted for the spreadsheet
worksheet- worksheet that will be written to
start_row- row to start reading from the out_write_sheet
qty_start- column where the qty columns start
qty_end- column where the qty columns end
header_row- row where the headers are located in the out_write_sheet
sheet_end- row that the out_write_sheet has no data
range1- first range of cells to multiply with the
second range of cells
range2- second range of cells to multiply with the
first range of cells
sheet- sheet name in out_write_sheet
formula- formula to add to the total sheet
part_num- part number in the out_write_sheet
row- row of data to add to the total sheet
"""


def update_totals(row, part_num, config_dict, workbook, header_list):
    header_row = config_dict["total_header_row"]
    total_row = row - config_dict["header_row"] + config_dict["total_header_row"]

    title = config_dict["total_sheet_name"]
    sheet_headers = config_dict["total_sheet_headers"]

    # adds new sheet if the sheet hasn't been created
    if title not in workbook.sheetnames:
        workbook.create_sheet(title=title)
        worksheet = workbook[title]

        # add specified headers
        for row in range(1, len(sheet_headers) + 1):
            utils.add_header(worksheet.cell(header_row, row), sheet_headers[row - 1])

    worksheet = workbook[title]

    qty_start = config_dict["qty_start"] - 1
    qty_end = len(header_list) - 1
    header_row = config_dict["header_row"]

    range1 = utils.get_range(qty_start, qty_end, row, row)
    range2 = utils.get_range(qty_start, qty_end, header_row - 1, header_row - 1)
    sheet = config_dict["out_sheet_name"] + "!"
    formula = "=IFERROR(SUMPRODUCT({0}{1}, {2}{3}),0)".format(sheet, range1, sheet, range2)

    worksheet.cell(total_row, 1).value = part_num
    worksheet.cell(total_row, 2).value = formula
